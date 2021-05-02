import openpyxl
import os

os.chdir('C:\\Users\\Saimun\\Python ardit\\reading excel xls')

workbook = openpyxl.load_workbook('example.xlsx')

# print(type(workbook))

print(workbook.get_sheet_names())

sheet = workbook.get_sheet_by_name('Sheet1')

cell1 = sheet['B1'].value

print(cell1)

cell2 = sheet['C1'].value
print(cell2)

# print(sheet.cell(row=1, column=2).value)

for i in range(1,8):
    print(sheet.cell(row=i, column=2).value)

for i in range(1,8):
    print(sheet.cell(row=i, column=3).value)